import uuid

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

import app.models as _models  # noqa: F401
from app.core.database import get_db
from app.core.security import hash_password
from app.main import app
from app.models.base import Base
from app.models.equipment import Equipment, EquipmentCategory, EquipmentStatus
from app.models.location import Location
from app.models.user import User, UserRole


@pytest.fixture(scope="function")
def test_db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)

    @event.listens_for(engine, "connect")
    def set_pragma(conn, _):
        conn.cursor().execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)

    def override():
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override
    db = SessionLocal()
    yield db
    db.close()
    app.dependency_overrides.clear()
    Base.metadata.drop_all(engine)


@pytest.fixture
def client(test_db):
    return TestClient(app)


@pytest.fixture
def director(test_db: Session) -> User:
    user = User(id=uuid.uuid4(), email="d@test.hu", full_name="D", role=UserRole.director,
                password_hash=hash_password("pass1234"), is_active=True)
    test_db.add(user)
    test_db.commit()
    return user


@pytest.fixture
def auth_headers(client, director):
    r = client.post("/api/v1/auth/login", json={"email": "d@test.hu", "password": "pass1234"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


@pytest.fixture
def location(test_db: Session) -> Location:
    loc = Location(id=uuid.uuid4(), name="Teszt Telephely")
    test_db.add(loc)
    test_db.commit()
    return loc


@pytest.fixture
def active_eq(test_db: Session, location: Location, director: User) -> Equipment:
    eq = Equipment(id=uuid.uuid4(), name="Laptop", category=EquipmentCategory.laptop,
                   qr_code=f"halo-inv://{uuid.uuid4()}", location_id=location.id,
                   created_by_id=director.id, status=EquipmentStatus.active)
    test_db.add(eq)
    test_db.commit()
    return eq


def test_retire_equipment(client, auth_headers, active_eq):
    resp = client.post(f"/api/v1/equipment/{active_eq.id}/retire",
                       json={"reason": "Teljesen elromlott"}, headers=auth_headers)
    assert resp.status_code == 200
    assert resp.json()["status"] == "retired"
    assert resp.json()["retirement_reason"] == "Teljesen elromlott"


def test_retire_already_retired(client, auth_headers, active_eq, test_db):
    # First retire
    client.post(f"/api/v1/equipment/{active_eq.id}/retire",
                json={"reason": "Elromlott ezután"}, headers=auth_headers)
    # Second retire → 409
    resp = client.post(f"/api/v1/equipment/{active_eq.id}/retire",
                       json={"reason": "Megint"}, headers=auth_headers)
    assert resp.status_code == 409


def test_retire_empty_reason(client, auth_headers, active_eq):
    resp = client.post(f"/api/v1/equipment/{active_eq.id}/retire",
                       json={"reason": "hi"}, headers=auth_headers)
    assert resp.status_code == 422


def test_qr_returns_png(client, auth_headers, active_eq):
    resp = client.get(f"/api/v1/equipment/{active_eq.id}/qr", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "image/png"
    # Check PNG magic bytes
    assert resp.content[:4] == b'\x89PNG'


def test_export_returns_xlsx(client, auth_headers, active_eq):
    resp = client.get("/api/v1/equipment/export", headers=auth_headers)
    assert resp.status_code == 200
    ct = resp.headers["content-type"]
    assert "spreadsheetml" in ct or "xlsx" in ct
    # Validate xlsx by re-reading with openpyxl
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 2  # header + at least 1 data row


def test_export_retired_only(client, auth_headers, active_eq, test_db, director, location):
    # Add a retired equipment
    retired = Equipment(id=uuid.uuid4(), name="Retired", category=EquipmentCategory.printer,
                        qr_code=f"halo-inv://{uuid.uuid4()}", location_id=location.id,
                        created_by_id=director.id, status=EquipmentStatus.retired,
                        retirement_reason="Elromlott")
    test_db.add(retired)
    test_db.commit()

    resp = client.get("/api/v1/equipment/export?status=retired", headers=auth_headers)
    assert resp.status_code == 200
    import io, openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Only header + retired row
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "Retired"
